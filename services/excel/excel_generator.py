from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO

from utils.format_states import FormatStates

from laboratory.domain.parameters import FQ_PARAMETERS, MICRO_PARAMETERS, UNIDADES, NORMA_CUMPLIMIENTO



class Excel:

    @staticmethod
    def generar_excel_reporte(muestras):

        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte Agua"

        titulo = Font(bold=True, size=13)
        encabezado = Font(bold=True)

        centro = Alignment(horizontal="center", wrap_text=True)
        izquierda = Alignment(horizontal="left", wrap_text=True)

        borde = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )
        

        fila = 1

        for datos in muestras:

            aplica_norma = Excel.aplica_norma(datos)

            # INFORMACION RECEPCION

            if aplica_norma:
                headers = [
                    "Parámetro",
                    "Unidad",
                    "Resultado",
                    "Resolución 2115/2007",
                    "Cumplimiento"
                ]
            else:
                headers = [
                    "Parámetro",
                    "Unidad",
                    "Resultado"
                ]

            num_cols = len(headers)
            ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=num_cols)
            ws.cell(row=fila, column=1, value="1. INFORMACIÓN DE RECEPCIÓN Y MUESTREO").font = titulo

            fila += 2

            info = [
                ("Código", datos.get("Código", "N/A"), "Fuente", datos.get("Fuente", "N/A")),
                ("Fecha Muestra", datos.get("Fecha", "N/A"), "Hora Muestra", datos.get("Hora", "N/A")),
                ("Fecha Recepción", datos.get("Fecha Recepción", "N/A"), "Recepcionista", datos.get("Recepcionó", "N/A")),
                ("Temp. Recepción", f"{datos.get('Temperatura Recepción', '0')} °C", "Tipo de Agua", datos.get("Tipo Agua", "N/A")),
                ("Estado FQ", FormatStates.format_estado(datos.get("Estado_FQ")), "Estado Micro", FormatStates.format_estado(datos.get("Estado_Micro"))),
                ("Estado del análisis", FormatStates.estado_analisis(datos), "", "")
            ]

            for campo1, valor1, campo2, valor2 in info:

                c1 = ws.cell(row=fila, column=1, value=campo1)
                c1.font = encabezado
                c1.alignment = izquierda

                c2 = ws.cell(row=fila, column=2, value=valor1)
                c2.alignment = izquierda

                if campo2:
                    c3 = ws.cell(row=fila, column=3, value=campo2)
                    c3.font = encabezado
                    c3.alignment = izquierda

                    c4 = ws.cell(row=fila, column=4, value=valor2)
                    c4.alignment = izquierda

                ws.row_dimensions[fila].height = 22
                fila += 1

            fila += 2

            # ANALISIS FQ
            ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=num_cols)
            ws.cell(row=fila, column=1, value="2. ANÁLISIS FÍSICO-QUÍMICO").font = titulo

            fila += 1

            for col, h in enumerate(headers, 1):
                c = ws.cell(row=fila, column=col, value=h)
                c.font = encabezado
                c.alignment = centro
                c.border = borde

            ws.row_dimensions[fila].height = 25
            fila += 1

            for parametro in FQ_PARAMETERS:
                resultado = FormatStates.resolver_resultado_fq(datos, parametro)

                unidad = UNIDADES.get(parametro, "")

                norma = "N/A"
                cumplimiento = "N/A"

                if aplica_norma:

                    cumple, valor_norma = Excel.cumple_limite(parametro, resultado)

                    norma = valor_norma

                    if cumple is None:
                        cumplimiento = "N/A"
                    else:
                        cumplimiento = "Cumple" if cumple else "No cumple"

                if aplica_norma:
                    fila_data = [
                        parametro,
                        unidad,
                        resultado,
                        norma,
                        cumplimiento
                    ]
                else:
                    fila_data = [
                        parametro,
                        unidad,
                        resultado
                    ]

                for col, val in enumerate(fila_data, 1):
                    c = ws.cell(row=fila, column=col, value=val)
                    c.border = borde
                    c.alignment = centro

                fila += 1

            fila += 2

            #ANALISIS MICRO

            ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=num_cols)
            ws.cell(row=fila, column=1, value="3. ANÁLISIS MICROBIOLÓGICO").font = titulo

            fila += 1

            for col, h in enumerate(headers, 1):
                c = ws.cell(row=fila, column=col, value=h)
                c.font = encabezado
                c.alignment = centro
                c.border = borde

            ws.row_dimensions[fila].height = 25
            fila += 1

            for parametro in MICRO_PARAMETERS:
                resultado = FormatStates.resolver_resultado_micro(datos, parametro)

                unidad = UNIDADES.get(parametro, "")

                norma = "N/A"
                cumplimiento = "N/A"

                if aplica_norma:

                    try:
                        valor_num = float(resultado)
                        cumple, valor_norma = Excel.cumple_limite(parametro, valor_num)

                        norma = valor_norma

                        if cumple is None:
                            cumplimiento = "N/A"
                        else:
                            cumplimiento = "Cumple" if cumple else "No cumple"

                    except:
                        norma = "N/A"
                        cumplimiento = "N/A"

                if aplica_norma:
                    fila_data = [
                        parametro,
                        unidad,
                        resultado,
                        norma,
                        cumplimiento
                    ]
                else:
                    fila_data = [
                        parametro,
                        unidad,
                        resultado
                    ]

                for col, val in enumerate(fila_data, 1):
                    c = ws.cell(row=fila, column=col, value=val)
                    c.border = borde
                    c.alignment = centro

                fila += 1

            fila += 4

        #Ajuste de las columnas
        for column in ws.columns:

            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))

            ws.column_dimensions[column_letter].width = max_length + 4

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        return buffer.getvalue()
    
    @staticmethod
    def aplica_norma(muestra):
        return (
            muestra.get("Tipo Agua") == "Agua potable (AP)"
            or muestra.get("Tipo Agua") == "Agua superficial (ASP)"
        )


    @staticmethod
    def cumple_limite(parametro, valor):

        regla = NORMA_CUMPLIMIENTO.get(parametro)

        if not regla:
            return None, "N/A"

        try:
            valor = float(valor)
        except:
            return None, "N/A"

        tipo = regla.get("type")

        if tipo == "range":

            minimo = regla.get("min")
            maximo = regla.get("max")

            if minimo is None or maximo is None:
                return None, "N/A"

            return minimo <= valor <= maximo, f"{minimo} - {maximo}"

        if tipo == "exact":

            esperado = regla.get("value")

            if esperado is None:
                return None, "N/A"

            return valor == esperado, str(esperado)

        return None, "N/A"